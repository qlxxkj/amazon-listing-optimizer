# export/export_to_autopart_template_v11.py
from openpyxl import load_workbook
from datetime import datetime
from db.save_data import get_all_data
import os, re, difflib, random

# ======================================================
# 🚗 一、字段映射定义（Amazon采集模板 → AUTO_PART 上传模板）
# ======================================================
FIELD_MAPPING = {
    "optimized_title": "Item Name",
    "optimized_description": "Product Description",
    "optimized_features": "Bullet Point",
    "search_keywords": "Item Type Keyword",
    "main_image": "Main Image URL",
}

# ======================================================
# 🧩 二、固定字段默认值
# ======================================================
FINAL_DEFAULTS = {
    "Listing Action": "Create or Replace (Full Update)",
    "Product Type": "AUTO_PART",
    "Brand Name": "YvuLDeQP",
    "Manufacturer": "YvuLDeQP",
    "Product ID Type": "EAN",
    "Fulfillment Channel Code (US)": "DEFAULT",
    "Item Condition": "New",
    "Quantity (US)": 100,
    "Handling Time (US)": 5,
    "Merchant Shipping Group (US)": "Migrated Template",
    "Number of Items": 1,
    "Exterior Finish": "Polished",
    "Required Assembly": "No",
    "Included Components": "Not Applicable",
    "Automotive Fit Type": "Universal Fit",
    "Country of Origin": "China",
    "Warranty Description": "Not Applicable",
    "Dangerous Goods Regulations": "Not Applicable",
    "Contains Liquid Contents?": "No",
    "Product Compliance Certificate": "Not Applicable",
    "Package Level": "Unit",
}


# ======================================================
# 🧮 三、唯一编码生成
# ======================================================
def generate_random_code(prefix_len=3, num_len=4):
    """随机生成SKU/型号"""
    letters = "".join(random.choice("ABCDEFGHIJKLMNOPQRSTUVWXYZ") for _ in range(prefix_len))
    numbers = "".join(random.choice("0123456789") for _ in range(num_len))
    return f"{letters}{numbers}"


def generate_unique_ean(existing_set):
    """生成符合 EAN13 校验规则的唯一编码"""
    while True:
        base = [random.randint(0, 9) for _ in range(12)]
        checksum = (10 - (sum(base[i] if i % 2 == 0 else base[i] * 3 for i in range(12)) % 10)) % 10
        ean = "".join(map(str, base + [checksum]))
        if ean not in existing_set:
            existing_set.add(ean)
            return ean


# ======================================================
# 🧰 四、工具函数
# ======================================================
def normalize_header(name):
    if not name:
        return ""
    s = str(name)
    s = re.sub(r"[\u200b\u200c\u200d\uFEFF\n\r\t]", "", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip().lower()


def match_column(expected_name, headers):
    """模糊匹配字段名到模板列"""
    normalized_expected = normalize_header(expected_name)
    best_match = difflib.get_close_matches(normalized_expected, list(headers.keys()), n=1, cutoff=0.4)
    return headers[best_match[0]][0] if best_match else None


def safe_value(value):
    if isinstance(value, list):
        return "\n".join(str(v) for v in value)
    return "" if value is None else str(value)


def capitalize_unit(unit):
    """单位首字母大写"""
    if not unit:
        return ""
    return unit.strip().capitalize()


# ======================================================
# 🚀 五、主导出逻辑
# ======================================================
def export_to_autopart_template(start_date=None, end_date=None, template_path="templates/AUTO_PART.xlsm"):
    data_list = get_all_data(start_date=start_date, end_date=end_date)
    if not data_list:
        print("[EXPORT] 没有符合条件的数据")
        return None

    if not os.path.exists(template_path):
        raise FileNotFoundError(f"模板文件不存在: {template_path}")

    wb = load_workbook(template_path, keep_vba=True)
    ws = wb["Template"]

    # 查找表头
    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and any(str(c).strip().lower() == "sku" for c in row if c):
            header_row_idx = i
            break
    if not header_row_idx:
        raise ValueError("Cannot find the header row (row containing 'SKU')")

    headers = {}
    for idx, cell in enumerate(ws[header_row_idx], start=1):
        key = normalize_header(cell.value)
        headers.setdefault(key, []).append(idx)

    print(f"[EXPORT] 模板表头加载完成，共 {len(headers)} 列")

    # 检测提示行
    start_row = header_row_idx + 1
    for row_cells in ws.iter_rows(min_row=header_row_idx + 1, max_row=header_row_idx + 10):
        for cell in row_cells:
            if cell.value and isinstance(cell.value, str) and "prefilled" in cell.value.lower():
                start_row = cell.row + 1
                print(f"✅ 检测到提示行在第 {cell.row} 行，数据从第 {start_row} 行开始写入")
                break
        else:
            continue
        break

    used_eans = set()
    row = start_row

    # ======================================================
    # 🧾 数据写入
    # ======================================================
    for item in data_list:
        item = item._mapping
        cleaned = item.get("cleaned") or {}
        optimized = item.get("optimized") or {}

        # === 五点 ===
        features = optimized.get("optimized_features") or cleaned.get("features", [])

        # === SKU,型号，EAN，OEM Num,等唯一编码生成 ===
        sku = generate_random_code()
        model_num = generate_random_code()
        model_name = generate_random_code()
        part_num = generate_random_code()
        ean = generate_unique_ean(used_eans)
        OEM = cleaned.get("OEM_Part_Number", "")

        # === 价格 ===
        price_rate = 2.0 # 价格倍数 或者汇率
        base_price = cleaned.get("price_value") or 0.0
        price = round(base_price * price_rate, 2)
        list_price = price
        your_price = str(price)

        # === 尺寸和重量 ===
        length, width, height = cleaned.get("item_length"), cleaned.get("item_width"), cleaned.get("item_height")
        unit = capitalize_unit(cleaned.get("item_size_unit"))
        weight_val, weight_unit = cleaned.get("item_weight_value"), capitalize_unit(cleaned.get("item_weight_unit"))

        variants = cleaned.get("variants", [])
        has_variants = len(variants) > 1
        parent_sku = sku + "_P" if has_variants else None

        # ✅ 父体
        if has_variants:
            ws.cell(row=row, column=match_column("SKU", headers), value=parent_sku)
            ws.cell(row=row, column=match_column("Parentage Level", headers), value="Parent")
            ws.cell(row=row, column=match_column("Variation Theme Name", headers), value="Color/Size")
            row += 1

        # ✅ 子体或单品
        targets = variants if has_variants else [{"main_image": cleaned.get("main_image"), "images": cleaned.get("other_images", [])}]
        for v in targets:
            child_sku = generate_random_code()
            ws.cell(row=row, column=match_column("SKU", headers), value=child_sku)

            if has_variants:
                ws.cell(row=row, column=match_column("Parent SKU", headers), value=parent_sku)
                ws.cell(row=row, column=match_column("Parentage Level", headers), value="Child")
                ws.cell(row=row, column=match_column("Variation Theme Name", headers), value="Color/Size")

            # 主字段
            for field, col_name in FIELD_MAPPING.items():
                col_idx = match_column(col_name, headers)
                if col_idx:
                    val = optimized.get(field, "")
                    ws.cell(row=row, column=col_idx, value=safe_value(val))

            # 默认字段
            for field, val in FINAL_DEFAULTS.items():
                col_idx = match_column(field, headers)
                if col_idx:
                    ws.cell(row=row, column=col_idx, value=val)

            # Bullet Point 分布
            if "bullet point" in headers:
                bullet_cols = headers["bullet point"]
                for i, col_idx in enumerate(bullet_cols):
                    if i < len(features):
                        ws.cell(row=row, column=col_idx, value=features[i])

            # 额外字段
            extra_fields = {
                "List Price": list_price,
                "Your Price USD (Sell on Amazon, US)": your_price,
                "Item Length": length,
                "Item Width": width,
                "Item Height": height,
                "Item Length Unit": unit,
                "Item Width Unit": unit,
                "Item Height Unit": unit,
                "Item Weight": weight_val,
                "Item Weight Unit": weight_unit,
                "Item Package Length": length,
                "Item Package Width": width,
                "Item Package Height": height,
                "Package Length Unit": unit,
                "Package Width Unit": unit,
                "Package Height Unit": unit,
                "Package Weight": weight_val,
                "Package Weight Unit": weight_unit,
                "External Product ID": ean,
                "Model Number": model_num,
                "Model Name": model_name,
                "Part Number": part_num,
                "OEM Equivalent Part Number": OEM,
            }

            for field, val in extra_fields.items():
                col_idx = match_column(field, headers)
                if col_idx:
                    ws.cell(row=row, column=col_idx, value=val)

            # 图片写入
            main_img = v.get("main_image", "")
            other_imgs = v.get("images", []) or cleaned.get("other_images", [])
            if match_column("Main Image URL", headers):
                ws.cell(row=row, column=match_column("Main Image URL", headers), value=main_img)
            if "other image url" in headers:
                for i, col_idx in enumerate(headers["other image url"]):
                    if i < len(other_imgs):
                        ws.cell(row=row, column=col_idx, value=other_imgs[i])

            row += 1

    # ======================================================
    # 🧾 导出保存
    # ======================================================
    os.makedirs("out", exist_ok=True)
    out_path = f"out/auto_part_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsm"
    wb.save(out_path)
    print(f"[EXPORT] ✅ 导出完成: {out_path}")
    return out_path
